import streamlit as st
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# Page setup
# ============================================================

st.set_page_config(
    page_title="Field Visit Compliance",
    page_icon="🧭",
    layout="wide",
)

# ============================================================
# Visual identity
# ============================================================

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Sora:wght@600;700;800&family=Inter:wght@400;500;600&family=JetBrains+Mono:wght@500;600&display=swap');

:root {
    --navy: #12283B;
    --navy-2: #1C3A52;
    --bg: #F5F7F8;
    --card: #FFFFFF;
    --border: #E1E6EA;
    --ink: #1B2733;
    --muted: #64707A;
    --teal: #1C9C8B;
    --teal-bg: #E1F5F1;
    --amber: #E8A33D;
    --amber-bg: #FCF1DF;
    --coral: #E2584A;
    --coral-bg: #FBE7E4;
}

html, body, [class*="css"]  {
    font-family: 'Inter', sans-serif;
    color: var(--ink);
}

.stApp {
    background: var(--bg);
}

#MainMenu, footer, header {visibility: hidden;}

.block-container {
    padding-top: 2rem;
    max-width: 1180px;
}

/* ---------- Hero ---------- */

.hero {
    background: linear-gradient(135deg, var(--navy) 0%, var(--navy-2) 100%);
    border-radius: 18px;
    padding: 2.4rem 2.6rem;
    margin-bottom: 1.8rem;
    position: relative;
    overflow: hidden;
}

.hero-eyebrow {
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.75rem;
    letter-spacing: 0.14em;
    color: #7FD9CC;
    text-transform: uppercase;
    margin-bottom: 0.6rem;
}

.hero h1 {
    font-family: 'Sora', sans-serif;
    font-weight: 800;
    font-size: 2.1rem;
    color: #FFFFFF;
    margin: 0 0 0.5rem 0;
    letter-spacing: -0.02em;
}

.hero p {
    color: #B9C6D2;
    font-size: 0.98rem;
    margin: 0;
    max-width: 640px;
}

.route-line {
    height: 2px;
    margin-top: 1.6rem;
    background: repeating-linear-gradient(to right, #7FD9CC 0 9px, transparent 9px 18px);
    position: relative;
}
.route-line::before, .route-line::after {
    content: '';
    position: absolute;
    top: 50%;
    transform: translateY(-50%);
    width: 8px; height: 8px;
    border-radius: 50%;
    background: #7FD9CC;
}
.route-line::before { left: -4px; }
.route-line::after { right: -4px; }

/* ---------- Section divider (signature element) ---------- */

.checkpoint-divider {
    height: 2px;
    margin: 2.2rem 0;
    background: repeating-linear-gradient(to right, var(--teal) 0 9px, transparent 9px 20px);
    position: relative;
}
.checkpoint-divider::before, .checkpoint-divider::after {
    content: '';
    position: absolute;
    top: 50%;
    transform: translateY(-50%);
    width: 9px; height: 9px;
    border-radius: 50%;
    background: var(--navy);
}
.checkpoint-divider::before { left: -4px; }
.checkpoint-divider::after { right: -4px; }

/* ---------- Cards ---------- */

.card {
    background: var(--card);
    border: 1px solid var(--border);
    border-radius: 14px;
    padding: 1.4rem 1.6rem;
}

.section-label {
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.72rem;
    letter-spacing: 0.12em;
    text-transform: uppercase;
    color: var(--muted);
    margin-bottom: 0.6rem;
}

/* ---------- KPI cards ---------- */

.kpi-grid {
    display: grid;
    grid-template-columns: repeat(4, 1fr);
    gap: 0.9rem;
    margin-bottom: 0.4rem;
}

.kpi-card {
    background: var(--card);
    border: 1px solid var(--border);
    border-left: 4px solid var(--navy);
    border-radius: 12px;
    padding: 1.1rem 1.3rem;
}

.kpi-card.total { border-left-color: var(--navy); }
.kpi-card.full { border-left-color: var(--teal); }
.kpi-card.partial { border-left-color: var(--amber); }
.kpi-card.non { border-left-color: var(--coral); }

.kpi-label {
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.72rem;
    letter-spacing: 0.08em;
    text-transform: uppercase;
    color: var(--muted);
    margin-bottom: 0.35rem;
}

.kpi-value {
    font-family: 'Sora', sans-serif;
    font-weight: 700;
    font-size: 1.9rem;
    color: var(--ink);
    line-height: 1;
}

.kpi-sub {
    font-family: 'JetBrains Mono', monospace;
    font-size: 0.78rem;
    color: var(--muted);
    margin-top: 0.3rem;
}

/* ---------- Uploader ---------- */

[data-testid="stFileUploaderDropzone"] {
    background: var(--card);
    border: 1.5px dashed #C4CDD3;
    border-radius: 14px;
}

/* ---------- Buttons ---------- */

.stDownloadButton button, .stButton button {
    background: var(--navy);
    color: #FFFFFF;
    border: none;
    border-radius: 10px;
    padding: 0.6rem 1.4rem;
    font-family: 'Inter', sans-serif;
    font-weight: 600;
    letter-spacing: 0.01em;
    transition: background 0.15s ease;
}
.stDownloadButton button:hover, .stButton button:hover {
    background: var(--teal);
    color: #FFFFFF;
}

/* ---------- Dataframe ---------- */

[data-testid="stDataFrame"] {
    border: 1px solid var(--border);
    border-radius: 12px;
    overflow: hidden;
}

</style>
""", unsafe_allow_html=True)

# ============================================================
# Hero
# ============================================================

st.markdown("""
<div class="hero">
    <div class="hero-eyebrow">Territory &middot; Cluster &middot; Field Visits</div>
    <h1>🧭 Field Visit Compliance</h1>
    <p>Upload the daily CRM export to check first-half and second-half field time against
    target, spot bunched check-ins, and export a formatted report for review.</p>
    <div class="route-line"></div>
</div>
""", unsafe_allow_html=True)

# ============================================================
# Upload
# ============================================================

uploaded_file = st.file_uploader(
    "Upload CRM Excel (.xlsx)",
    type=["xlsx"],
    help="Expects columns: Territory, Cluster, EmployeeName, City, LoginDate"
)

if not uploaded_file:
    st.markdown("""
    <div class="card">
        <div class="section-label">Waiting for a file</div>
        <p style="color:var(--muted); margin:0;">
        Drop today's CRM export above to generate the compliance summary.
        Field-half (FH) visits before 2 PM need at least 2 hours in-field;
        second-half (SH) visits need at least 3 hours.
        </p>
    </div>
    """, unsafe_allow_html=True)

if uploaded_file:

    # Read Excel from the uploaded file (not a hardcoded local path)
    df = pd.read_excel(uploaded_file)

    # Convert LoginDate
    df["LoginDate"] = pd.to_datetime(
        df["LoginDate"],
        errors="coerce"
    )

    # Remove blank dates
    df = df.dropna(subset=["LoginDate"])

    # FH before 2 PM
    df["Session"] = df["LoginDate"].dt.hour.apply(
        lambda x: "FH" if x < 14 else "SH"
    )

    summary = []

    group_cols = [
        "Territory",
        "Cluster",
        "EmployeeName"
    ]

    for key, grp in df.groupby(group_cols):

        territory, cluster, emp = key

        fh = grp[grp["Session"] == "FH"]
        sh = grp[grp["Session"] == "SH"]

        # ---------------- FH ----------------

        fh_locations = ", ".join(
            fh["City"].dropna().astype(str).unique()
        )

        fh_visits = len(fh)

        if len(fh):
            fh_first = fh["LoginDate"].min()
            fh_last = fh["LoginDate"].max()

            fh_minutes = int(
                (fh_last - fh_first).total_seconds() / 60
            )

            fh_duration = f"{fh_minutes//60}:{fh_minutes%60:02d}"
            fh_status = "Met" if fh_minutes >= 120 else "Not Met"
        else:
            fh_first = pd.NaT
            fh_last = pd.NaT
            fh_duration = "00:00"
            fh_status = "Not Met"

        # ---------------- SH ----------------

        sh_locations = ", ".join(
            sh["City"].dropna().astype(str).unique()
        )

        sh_visits = len(sh)

        if len(sh):
            sh_first = sh["LoginDate"].min()
            sh_last = sh["LoginDate"].max()

            sh_minutes = int(
                (sh_last - sh_first).total_seconds() / 60
            )

            sh_duration = f"{sh_minutes//60}:{sh_minutes%60:02d}"
            sh_status = "Met" if sh_minutes >= 180 else "Not Met"
        else:
            sh_first = pd.NaT
            sh_last = pd.NaT
            sh_duration = "00:00"
            sh_status = "Not Met"

        # ---------------- Total ----------------

        total = fh_visits + sh_visits

        if fh_status == "Met" and sh_status == "Met":
            compliance = "Fully Compliant"
        elif fh_status == "Met" or sh_status == "Met":
            compliance = "Partially Compliant"
        else:
            compliance = "Non-Compliant"

        # ---------------- Remarks ----------------

        remarks_list = []

        if len(fh) > 1:
            fh_sorted = fh.sort_values("LoginDate")
            fh_gaps = (
                fh_sorted["LoginDate"]
                .diff()
                .dt.total_seconds()
                .dropna() / 60
            )
            if fh_gaps.mean() < 5:
                remarks_list.append(
                    "FH visits bunched (avg gap <5m); verify field presence."
                )

        if len(sh) > 1:
            sh_sorted = sh.sort_values("LoginDate")
            sh_gaps = (
                sh_sorted["LoginDate"]
                .diff()
                .dt.total_seconds()
                .dropna() / 60
            )
            if sh_gaps.mean() < 5:
                remarks_list.append(
                    "SH visits bunched (avg gap <5m); verify field presence."
                )

        remarks = " ".join(remarks_list)

        summary.append([
            territory,
            cluster,
            emp,
            fh_locations,
            fh_visits,
            fh_first,
            fh_last,
            fh_duration,
            fh_status,
            sh_locations,
            sh_visits,
            sh_first,
            sh_last,
            sh_duration,
            sh_status,
            total,
            compliance,
            remarks
        ])

    cols = [
        "Territory",
        "Cluster",
        "CM Name",
        "FH Locations Visited",
        "FH Visits",
        "FH First Visit",
        "FH Last Visit",
        "FH Duration",
        "FH Status",
        "SH Locations Visited",
        "SH Visits",
        "SH First Visit",
        "SH Last Visit",
        "SH Duration",
        "SH Status",
        "Total Visits (Day)",
        "Overall Compliance",
        "Remarks"
    ]

    summary_df = pd.DataFrame(summary, columns=cols)
    summary_df.insert(0, "S.No", range(1, len(summary_df) + 1))

    # Show only time
    time_cols = [
        "FH First Visit",
        "FH Last Visit",
        "SH First Visit",
        "SH Last Visit"
    ]

    for col in time_cols:
        summary_df[col] = pd.to_datetime(
            summary_df[col],
            errors="coerce"
        ).dt.strftime("%H:%M").fillna("")

    # ============================================================
    # KPI strip
    # ============================================================

    total_cms = len(summary_df)
    full_count = int((summary_df["Overall Compliance"] == "Fully Compliant").sum())
    partial_count = int((summary_df["Overall Compliance"] == "Partially Compliant").sum())
    non_count = int((summary_df["Overall Compliance"] == "Non-Compliant").sum())

    def pct(n):
        return f"{(n / total_cms * 100):.0f}%" if total_cms else "0%"

    st.markdown(f"""
    <div class="kpi-grid">
        <div class="kpi-card total">
            <div class="kpi-label">CMs Tracked</div>
            <div class="kpi-value">{total_cms}</div>
            <div class="kpi-sub">across all territories</div>
        </div>
        <div class="kpi-card full">
            <div class="kpi-label">Fully Compliant</div>
            <div class="kpi-value">{full_count}</div>
            <div class="kpi-sub">{pct(full_count)} of CMs</div>
        </div>
        <div class="kpi-card partial">
            <div class="kpi-label">Partially Compliant</div>
            <div class="kpi-value">{partial_count}</div>
            <div class="kpi-sub">{pct(partial_count)} of CMs</div>
        </div>
        <div class="kpi-card non">
            <div class="kpi-label">Non-Compliant</div>
            <div class="kpi-value">{non_count}</div>
            <div class="kpi-sub">{pct(non_count)} of CMs</div>
        </div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown('<div class="checkpoint-divider"></div>', unsafe_allow_html=True)

    # ============================================================
    # Table preview
    # ============================================================

    st.markdown('<div class="section-label">Summary — one row per CM</div>', unsafe_allow_html=True)

    def style_status(val):
        if val in ("Met", "Fully Compliant"):
            return "background-color:#E1F5F1; color:#127365; font-weight:600;"
        if val == "Partially Compliant":
            return "background-color:#FCF1DF; color:#9C6A17; font-weight:600;"
        if val in ("Not Met", "Non-Compliant"):
            return "background-color:#FBE7E4; color:#B03A2E; font-weight:600;"
        return ""

    try:
        styled = summary_df.style.applymap(
            style_status,
            subset=["FH Status", "SH Status", "Overall Compliance"]
        )
        st.dataframe(styled, use_container_width=True, hide_index=True)
    except Exception:
        st.dataframe(summary_df, use_container_width=True, hide_index=True)

    # ============================================================
    # Write formatted Excel to memory (single source of truth for download)
    # ============================================================

    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df.to_excel(writer, index=False, sheet_name="Summary")

    output.seek(0)

    # Reload the in-memory workbook to apply formatting
    wb = load_workbook(output)
    ws = wb["Summary"]

    # ==========================
    # Colors
    # ==========================

    header_fill = PatternFill(fill_type="solid", fgColor="1F4E78")   # Dark Blue
    green_fill = PatternFill(fill_type="solid", fgColor="C6EFCE")
    red_fill = PatternFill(fill_type="solid", fgColor="FFC7CE")
    yellow_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")

    header_font = Font(color="FFFFFF", bold=True)

    thin = Side(border_style="thin", color="D9D9D9")

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin
    )

    # ==========================
    # Header Formatting
    # ==========================

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        cell.border = border

    # ==========================
    # Column Numbers
    # ==========================

    headers = {}
    for cell in ws[1]:
        headers[cell.value] = cell.column

    fh_status_col = headers["FH Status"]
    sh_status_col = headers["SH Status"]
    overall_col = headers["Overall Compliance"]

    # ==========================
    # Row Formatting
    # ==========================

    for row in range(2, ws.max_row + 1):

        for col in range(1, ws.max_column + 1):
            c = ws.cell(row=row, column=col)
            c.border = border
            c.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

        c = ws.cell(row=row, column=fh_status_col)
        c.fill = green_fill if c.value == "Met" else red_fill

        c = ws.cell(row=row, column=sh_status_col)
        c.fill = green_fill if c.value == "Met" else red_fill

        c = ws.cell(row=row, column=overall_col)
        if c.value == "Fully Compliant":
            c.fill = green_fill
        elif c.value == "Partially Compliant":
            c.fill = yellow_fill
        elif c.value == "Non-Compliant":
            c.fill = red_fill

    # ==========================
    # Auto Width
    # ==========================

    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except Exception:
                pass

        ws.column_dimensions[column_letter].width = min(max_length + 3, 40)

    # ==========================
    # Freeze Header / Auto Filter / Row Height
    # ==========================

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 35

    # ==========================
    # Save formatted workbook back to memory for download
    # ==========================

    final_output = BytesIO()
    wb.save(final_output)
    final_output.seek(0)

    st.markdown('<div class="checkpoint-divider"></div>', unsafe_allow_html=True)

    st.markdown("""
    <div class="card">
        <div class="section-label">Export</div>
        <p style="color:var(--muted); margin:0 0 1rem 0;">
        Download the same summary as a formatted, ready-to-share Excel report.
        </p>
    </div>
    """, unsafe_allow_html=True)

    st.download_button(
        "Download summary report (.xlsx)",
        data=final_output,
        file_name="Summary_Report.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )